from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from apps.beneficiario.models import *
from openpyxl import Workbook
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from weasyprint import HTML
from django.core.files.storage import FileSystemStorage
def informes(request):
    return render(request,'informes.html')

def infoSisben(request):
    if request.method == 'POST':
        fi = request.POST['fechaInicial']
        ff = request.POST['fechaFinal']
        beneficiarios = Carnet.objects.filter(fechaExpide__range=[fi,ff])
        if beneficiarios:
            fin = parse_date(fi)
            ffn = parse_date(ff)
            tl = beneficiarios.count
            return render(request,'infoSisben.html',{ 'beneficiarios':beneficiarios,'fi':fin,'ff':ffn,'tl':tl, 'dan':fi, 'dfn':ff })
        else:
            messages.error(request,'No se encontraron registros')
            return redirect('/informes/infoSisben')
    return render(request,'infoSisben.html')

def infoSisbenPdf(request):
    fi = request.GET['fi']
    ff = request.GET['ff']

    beneficiarios = Carnet.objects.filter(fechaExpide__range=[fi,ff])
    fin = parse_date(fi)
    ffn = parse_date(ff)
    tl = beneficiarios.count
    html_string = render_to_string('infoSisbenPdf.html',{'beneficiarios':beneficiarios,'fi':fin,'ff':ffn,'tl':tl })
    html = HTML(string=html_string)
    html.write_pdf(target='/tmp/infoSisben.pdf')

    fs = FileSystemStorage('/tmp')
    with fs.open('infoSisben.pdf') as pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'filename="infoSisben.pdf"'.format("order.id")
        return response

def informeVca(request):
    beneficiarios = Beneficiario.objects.all()
    wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda

        #Creamos los encabezados desde la celda B3 hasta la E3
    ws['A1'] = 'Enfoque Diferencial'
    ws.merge_cells('A1:D1')
    ws['E1'] = '0 a 5 años'
    ws['F1'] = '6 a 11 años'
    ws['G1'] = '12 a 17 años'
    ws['H1'] = '18 a 28 años'
    ws['I1'] = '29 a 60 años'
    ws['J1'] = '61 años o mas'
    ws['K1'] = 'No disponible'
    ws['L1'] = 'Total por grupo'
    ws['A2'] = 'Total Mujeres'
    ws.merge_cells('A2:D2')
    ws['A3'] = 'Total Hombres'
    ws.merge_cells('A3:D3')
    beneficiario = Beneficiario.objects.all()
    hombres = beneficiario.filter(genero='H')
    mujeres = beneficiario.filter(genero='M')

    r1 = []
    r2 = []
    r3 = []
    r4 = []
    r5 = []
    r6 = []
    #import pdb; pdb.set_trace()
    for y in mujeres:
        if y.getEdad() >= 0 and y.getEdad() <= 5:
            r1.append(y)
    for y in mujeres:
        if y.getEdad() >= 6 and y.getEdad() <= 11:
            r2.append(y)
    for y in mujeres:
        if y.getEdad() >= 12 and y.getEdad() <= 17:
            r3.append(y)
    for y in mujeres:
        if y.getEdad() >= 18 and y.getEdad() <= 28:
            r4.append(y)
    for y in mujeres:
        if y.getEdad() >= 29 and y.getEdad() <= 60:
            r5.append(y)
    for y in mujeres:
        if y.getEdad() >= 61:
            r6.append(y)



    #ws.cell(row=2,column=5).value = beneficiario.filter(genero='M').count()
    ws.cell(row=2,column=5).value = len(r1)
    ws.cell(row=2,column=6).value = len(r2)
    ws.cell(row=2,column=7).value = len(r3)
    ws.cell(row=2,column=8).value = len(r4)
    ws.cell(row=2,column=9).value = len(r5)
    ws.cell(row=2,column=10).value = len(r6)
    ws.cell(row=3,column=5).value = beneficiario.filter(genero='H').count()

        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas

        #Establecemos el nombre del archivo
    nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
